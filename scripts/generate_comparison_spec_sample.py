#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
고속차량 제원 비교표 롱 포맷 변환 테스트용 샘플 xlsx 생성.

실행 (프로젝트 루트):
  python scripts/generate_comparison_spec_sample.py

출력:
  data/samples/고속차량 제원 비교표_파싱샘플.xlsx

검증:
  python -c "import sys; sys.path.insert(0,'src'); from pathlib import Path; from comparison_spec_long_format import transform_comparison_spec_workbook; from config import TEMP_DIR; p=Path('data/samples/고속차량 제원 비교표_파싱샘플.xlsx'); print(transform_comparison_spec_workbook(p, TEMP_DIR))"
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import openpyxl
from openpyxl.styles import Alignment


def main() -> None:
    out_dir = ROOT / "data" / "samples"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "고속차량 제원 비교표_파싱샘플.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "제원비교"

    # 1~4행: 제목·여백 (실무 파일과 유사)
    ws.merge_cells("A1:H1")
    ws["A1"] = "고속차량 제원"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 5행: 헤더 — 구분 | 상세내역 | 차량열…
    hr = 5
    ws.cell(row=hr, column=1, value="구분")
    ws.cell(row=hr, column=2, value="상세내역")
    ws.cell(row=hr, column=3, value="KTX")
    ws.cell(row=hr, column=4, value="KTX-산천 I")
    ws.cell(row=hr, column=5, value="SRT")

    # 일반사항 (A6:A7 병합)
    ws.merge_cells(start_row=hr + 1, start_column=1, end_row=hr + 2, end_column=1)
    ws.cell(row=hr + 1, column=1, value="일반사항")
    ws.cell(row=hr + 1, column=2, value="도입년도")
    ws.cell(row=hr + 1, column=3, value="2004년")
    ws.cell(row=hr + 1, column=4, value="2010~2012년")
    ws.cell(row=hr + 1, column=5, value="2016년")
    ws.cell(row=hr + 2, column=2, value="도입량수")
    ws.cell(row=hr + 2, column=3, value="920량")
    ws.cell(row=hr + 2, column=4, value="240량")
    ws.cell(row=hr + 2, column=5, value="400량")

    # 차량일반 (A8:A9 병합)
    ws.merge_cells(start_row=hr + 3, start_column=1, end_row=hr + 4, end_column=1)
    ws.cell(row=hr + 3, column=1, value="차량일반")
    ws.cell(row=hr + 3, column=2, value="편성(량)")
    ws.cell(row=hr + 3, column=3, value="20량")
    ws.cell(row=hr + 3, column=4, value="10량")
    ws.cell(row=hr + 3, column=5, value="10량")
    ws.cell(row=hr + 4, column=2, value="편성길이")
    ws.cell(row=hr + 4, column=3, value="388 m")
    ws.cell(row=hr + 4, column=4, value="201 m")
    ws.cell(row=hr + 4, column=5, value="200 m")

    wb.save(path)
    print(f"샘플 저장: {path}")
    print(f"  시트: {ws.title}, 헤더 행: {hr}, 파일명 접두어: 고속차량 제원 비교표")


if __name__ == "__main__":
    main()
