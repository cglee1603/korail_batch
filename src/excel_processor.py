"""
엑셀 파일 처리 모듈
시트별 헤더 자동 감지, 하이퍼링크 추출, 숨김 행 제외, 시트 타입 감지
"""
from typing import List, Dict, Tuple, Optional, Any
from pathlib import Path
from enum import Enum
import re
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import Cell
from logger import logger
from config import (
    TEST_MODE, TEST_MAX_SHEETS, TEST_MAX_ROWS,
    SHEET_TYPE_KEYWORDS, COLUMN_NAME_MAPPINGS,
    MAX_TEXT_LENGTH, ROW_SEPARATOR, TEXT_ENCODING
)


class SheetType(Enum):
    """시트 타입 분류"""
    TOC = "목차"  # 목차 시트 (처리 안 함)
    REV_MANAGED = "REV관리"  # REV + WBS 컬럼 존재
    VERSION_MANAGED = "작성버전관리"  # 작성버전 + 관리번호 컬럼 존재
    ATTACHMENT = "첨부파일"  # 하이퍼링크만 존재
    HISTORY = "이력관리"  # 이력관리 시트
    SOFTWARE = "소프트웨어형상"  # 소프트웨어 형상기록 시트
    UNKNOWN = "미분류"  # 알 수 없는 타입


class ExcelProcessor:
    """엑셀 파일 처리 클래스"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        # data_only=True로 로드한 워크북(수식의 계산된 값 접근용, 지연 로드)
        self._workbook_data_only = None
        # 시트별 컬럼 숨김 캐시: { sheet_name: {col_idx: bool_hidden} }
        self._sheet_col_hidden_map: Dict[str, Dict[int, bool]] = {}
        
    def load_workbook(self):
        """엑셀 파일 로드"""
        try:
            self.workbook = openpyxl.load_workbook(
                self.excel_path, 
                data_only=False,  # 하이퍼링크 읽기 위해 False
                keep_vba=False
            )
            logger.info(f"엑셀 파일 로드 완료: {self.excel_path}")
            return True
        except Exception as e:
            logger.error(f"엑셀 파일 로드 실패: {e}")
            return False
    
    def _ensure_data_only_workbook(self):
        """
        data_only=True로 계산된 값을 담은 워크북을 지연 로드한다.
        - 수식 셀의 '값'을 읽고 싶을 때 사용 (Excel에서 계산 후 저장된 캐시 값이 있어야 함)
        """
        if self._workbook_data_only is not None:
            return
        try:
            self._workbook_data_only = openpyxl.load_workbook(
                self.excel_path,
                data_only=True,
                keep_vba=False
            )
            logger.debug("data_only 워크북 로드 완료 (수식 캐시 값 접근용)")
        except Exception as e:
            logger.debug(f"data_only 워크북 로드 실패: {e}")
            self._workbook_data_only = None
    
    def get_sheet_names(self) -> List[str]:
        """모든 시트 이름 반환"""
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def detect_header_row(self, sheet: Worksheet) -> Tuple[int, int]:
        """
        헤더 행 자동 감지 (개선된 로직)
        
        헤더의 특징:
        1. 여러 컬럼에 값이 채워져 있음 (최소 3개)
        2. 괄호나 의미있는 단어 포함 (예: "년도(1)", "제목", "구분")
        3. 단순 숫자나 짧은 텍스트만 있으면 제목일 가능성 높음
        4. 다음 행부터 규칙적인 데이터가 있음
        """
        max_search_row = min(15, sheet.max_row + 1)
        candidates = []
        
        for row_idx in range(1, max_search_row):
            # 숨김 행은 스킵
            if self.is_row_hidden(sheet, row_idx):
                logger.debug(f"{row_idx}행 점수: 스킵(숨김 행)")
                continue

            visible_values: List[str] = []
            visible_values_for_analysis: List[str] = []

            # 숨김 컬럼은 제외하고 값 수집
            sheet_max_col = sheet.max_column or 1
            for col_idx in range(1, sheet_max_col + 1):
                if self.is_col_hidden(sheet, col_idx):
                    continue
                val = self._get_merged_top_left_value(sheet, row_idx, col_idx)
                if val is None:
                    continue
                cell_str = str(val).strip()
                if not cell_str:
                    continue
                visible_values.append(cell_str)
                if len(visible_values_for_analysis) < 10:
                    visible_values_for_analysis.append(cell_str)

            # '목차로 되돌아가기'가 포함된 행은 헤더 후보에서 제외
            try:
                if any("목차로 되돌아가기" in v for v in visible_values_for_analysis) or \
                   any("목차로 되돌아가기" in v for v in visible_values):
                    logger.debug(f"{row_idx}행 점수: 스킵(목차로 되돌아가기 포함)")
                    continue
            except Exception:
                pass

            non_empty_count = len(visible_values)

            # 최소 3개 이상의 값이 있어야 함
            if non_empty_count < 3:
                logger.debug(f"{row_idx}행 점수: 스킵(비어있지 않은 셀 {non_empty_count}개 < 3)")
                continue

            # 헤더 점수 계산
            score = 0

            # 1. 비어있지 않은 셀이 많을수록 점수 증가
            score += non_empty_count
            # 2. 헤더 특성 분석 (숨김 컬럼 제외한 선두 10개 값 기준)
            for cell_str in visible_values_for_analysis:
                # 괄호가 있으면 헤더일 가능성 높음 (예: "년도(1)")
                if '(' in cell_str and ')' in cell_str:
                    score += 5

                # 단순 숫자 1~2자리면 제목일 가능성 높음 (감점)
                if cell_str.isdigit() and len(cell_str) <= 2:
                    score -= 3

                # 일반적인 헤더 키워드
                header_keywords = ['년도', '제목', '구분', '번호', '이름', '코드', '상태', 
                                 '날짜', '작성', '담당', '버전', 'WBS', '종별', '관리']
                if any(keyword in cell_str for keyword in header_keywords):
                    score += 3

                # 너무 긴 텍스트는 제목일 가능성 높음 (감점)
                if len(cell_str) > 30:
                    score -= 2

            # 3. 다음 (가시) 행에 데이터가 있는지 확인
            if row_idx < sheet.max_row:
                next_row_idx = row_idx + 1
                # 바로 다음 행이 숨김이면 '다음 행' 평가는 생략
                if not self.is_row_hidden(sheet, next_row_idx):
                    next_non_empty = 0
                    for col_idx in range(1, sheet_max_col + 1):
                        if self.is_col_hidden(sheet, col_idx):
                            continue
                        val = self._get_merged_top_left_value(sheet, next_row_idx, col_idx)
                        if val is None:
                            continue
                        if str(val).strip():
                            next_non_empty += 1
                    # 다음 행에도 비슷한 개수의 데이터가 있으면 헤더일 가능성 높음
                    if next_non_empty >= non_empty_count * 0.5:
                        score += 3

            candidates.append((row_idx, score, non_empty_count))
            logger.debug(f"{row_idx}행 점수: {score} (비어있지 않은 셀: {non_empty_count}개, 숨김 제외)")
        
        # 가장 높은 점수의 행을 헤더로 선택 후, 해당 행의 마지막 의미 있는 컬럼을 max_col로 산출
        def compute_max_col(row_idx: int) -> int:
            sheet_max_col = sheet.max_column or 1
            last_col = 1
            for c in range(sheet_max_col, 0, -1):
                if self.is_col_hidden(sheet, c):
                    continue
                val = self._get_merged_top_left_value(sheet, row_idx, c)
                norm = self._normalize_text(val)
                if norm and not self._starts_with_legend_marker(norm):
                    last_col = c
                    break
            return last_col

        if candidates:
            candidates.sort(key=lambda x: x[1], reverse=True)
            best_row, best_score, best_count = candidates[0]
            max_col = compute_max_col(best_row)
            logger.info(f"헤더 행 감지: {best_row}행 (점수: {best_score}, 비어있지 않은 셀: {best_count}개)")
            return best_row, max_col

        # 기본값: 1행
        logger.warning("헤더 행을 찾지 못했습니다. 1행을 헤더로 사용합니다.")
        return 1, (sheet.max_column or 1)
    
    # ----- 계층형/병합 헤더 지원 유틸 -----
    def _get_merged_top_left_value(self, sheet: Worksheet, row: int, col: int) -> Optional[str]:
        """
        해당 좌표가 병합영역이면 좌상단 셀 값을 반환, 아니면 현재 셀 값
        (수정: 병합 영역 내라도 현재 셀에 값이 있으면 그 값을 우선 반환 - 헤더 오버라이딩 지원)
        """
        cell = sheet.cell(row=row, column=col)
        # 1. 현재 셀에 값이 있으면 우선 사용 (병합된 영역 내 숨겨진 값 읽기 용도)
        if cell.value is not None:
            return str(cell.value)

        # 2. 값이 없으면 병합 정보 확인하여 좌상단 값 가져오기
        coord = cell.coordinate
        for mrange in sheet.merged_cells.ranges:
            if coord in mrange:
                top_left = sheet.cell(row=mrange.min_row, column=mrange.min_col)
                return None if top_left.value is None else str(top_left.value)
        
        return None
    
    def _get_merged_top_left_value_evaluated(self, sheet: Worksheet, row: int, col: int) -> Optional[str]:
        """
        병합 좌상단 기준으로 '계산된 값(data_only)'을 우선 반환한다.
        - data_only 워크북이 없거나 값이 없으면 기존 값으로 폴백.
        """
        try:
            self._ensure_data_only_workbook()
            if self._workbook_data_only is not None:
                d_sheet = self._workbook_data_only[sheet.title]
                d_cell = d_sheet.cell(row=row, column=col)
                
                # 1. 현재 셀 값 우선 확인 (하위 헤더 오버라이딩 지원)
                if d_cell.value is not None:
                    return str(d_cell.value)

                # 2. 병합 영역이면 좌상단 값 확인
                coord = d_cell.coordinate
                for mrange in d_sheet.merged_cells.ranges:
                    if coord in mrange:
                        tl = d_sheet.cell(row=mrange.min_row, column=mrange.min_col)
                        return None if tl.value is None else str(tl.value)
                        
        except Exception as e:
            logger.debug(f"_get_merged_top_left_value_evaluated 실패(row={row}, col={col}): {e}")
        # 폴백: 기존(수식 문자열 포함 가능)
        return self._get_merged_top_left_value(sheet, row, col)

    def _row_non_empty_count(self, sheet: Worksheet, row: int, max_col: int) -> int:
        count = 0
        for col in range(1, max_col + 1):
            if self.is_col_hidden(sheet, col):
                continue
            val = self._get_merged_top_left_value(sheet, row, col)
            norm = self._normalize_text(val)
            if norm and not self._starts_with_legend_marker(norm):
                count += 1
        return count

    def _row_has_downward_merge_from(self, sheet: Worksheet, row: int) -> bool:
        """row를 포함하면서 아래로 확장되는 병합이 있는지"""
        for mrange in sheet.merged_cells.ranges:
            if mrange.min_row <= row <= mrange.max_row and mrange.max_row > row:
                return True
        return False

    def _make_unique_headers(self, headers: List[str]) -> List[str]:
        seen: dict[str, int] = {}
        unique: List[str] = []
        for h in headers:
            key = h or "Column"
            if key not in seen:
                seen[key] = 1
                unique.append(key)
            else:
                seen[key] += 1
                unique.append(f"{key} ({seen[key]})")
        return unique

    def _normalize_text(self, text: Optional[str]) -> Optional[str]:
        if text is None:
            return None
        # 모든 공백류(줄바꿈, 탭 포함)를 단일 공백으로 정규화
        s = " ".join(str(text).split())
        return s if s else None

    def _starts_with_legend_marker(self, text: str) -> bool:
        """행/헤더 카운팅에서 제외할 범례/참고 표식 여부(선두 특수기호)
        대괄호 등 헤더 그룹 표시는 유지하기 위해 과도한 범위는 피하고, 대표 표식만 필터링.
        """
        if not text:
            return False
        first = text[0]
        legend_markers = set([
            '※', '◎', '★', '☆', '●', '○', '■', '□', '◇', '◆', '▲', '△', '▼', '▽',
            '▶', '▷', '▸', '▹', '•', '∙', '·', '–', '—'
        ])
        return first in legend_markers

    def _is_symbolic_token(self, text: str) -> bool:
        """
        짧고(<=12), 공백 없고, 대문자/숫자/특수기호(-._) 위주인 토큰인지 검사
        예: INV, TYPE1, FTM, SEGMENT 등
        """
        if not text:
            return False
        # 스마트쿼트 등 통일
        text = str(text).replace("’", "'")
        # 공백/개행/마침표는 라벨 내 구분자로 취급하여 제거 후 평가
        # 예: "TDCS DU2" -> "TDCSDU2", "REV. TAG" -> "REVTAG"
        condensed = re.sub(r"[ \t\r\n\.]", "", text)
        # "TYPE 1"처럼 알파벳 + 공백 + 숫자 패턴도 동일하게 축약
        if re.fullmatch(r"[A-Za-z]+\s+\d+", text):
            condensed = re.sub(r"\s+", "", text)
        # 허용 길이(공백/마침표 제거 기준)
        if len(condensed) > 12:
            return False
        # 허용 문자: 영문/숫자/._-'/ (공백, 마침표는 위에서 제거된 상태)
        if not re.fullmatch(r"[A-Za-z0-9._\-'/]+", condensed):
            return False
        # 숫자만은 제외
        if condensed.isdigit():
            return False
        # 날짜/시간 패턴 간단 배제
        if re.search(r"\d{4}-\d{2}-\d{2}", condensed) or re.search(r"\d{2}/\d{1,2}/\d{1,2}", condensed):
            return False
        # 대문자 비율(알파 기준) 체크
        letters = [c for c in condensed if c.isalpha()]
        if letters:
            upper_ratio = sum(1 for c in letters if c.isupper()) / len(letters)
            if upper_ratio < 0.6:
                return False
        return True

    def _is_likely_symbolic_subheader_row(self, sheet: Worksheet, row: int, max_col: int) -> bool:
        """
        다음 행이 상징적 토큰들로 구성된 '부-헤더' 라인인지 휴리스틱으로 판정
        - 비어있지 않은 셀 수: 2~(max_col의 1/2) 범위
        - 비어있지 않은 값 중 80% 이상이 상징 토큰
        """
        non_empty = []
        for c in range(1, max_col + 1):
            if self.is_col_hidden(sheet, c):
                continue
            val = self._get_merged_top_left_value(sheet, row, c)
            norm = self._normalize_text(val)
            if norm:
                non_empty.append(norm)
        count = len(non_empty)
        if count < 2:
            return False
        if count > max(3, max_col // 2):
            return False
        symbolic = sum(1 for v in non_empty if self._is_symbolic_token(v))
        logger.info(f"row {row} symbolic {symbolic} count {count}, result {symbolic >= max(1, int(0.6 * count))}")
        return symbolic >= max(1, int(0.6 * count))

    def build_headers_and_data_start(self, sheet: Worksheet) -> Tuple[List[str], int, Tuple[int, int]]:
        """
        병합/다단 헤더를 고려하여 헤더를 생성하고 데이터 시작 행을 반환
        
        Returns:
            (headers, data_start_row, (header_start_row, header_end_row))
        """
        base_header_row, max_col = self.detect_header_row(sheet)

        include_prev = False
        include_next = False

        if base_header_row > 1:
            prev_non_empty = self._row_non_empty_count(sheet, base_header_row - 1, max_col)
            include_prev = prev_non_empty >= 2

        header_start = base_header_row - 1 if include_prev else base_header_row
        
        # 헤더 아래쪽 확장: 병합이 이어지거나 부-헤더(상징 토큰) 행이면 계속 포함
        header_end = base_header_row
        max_extension = 3  # 최대 3행까지 추가 확장
        extension_count = 0
        
        while extension_count < max_extension:
            next_row = header_end + 1
            if next_row > (sheet.max_row or next_row):
                break
            
            should_extend = False
            # 1. 현재 행에서 아래로 확장되는 병합이 있는지
            if self._row_has_downward_merge_from(sheet, header_end):
                should_extend = True
            # 2. 또는 다음 행이 상징적 부-헤더 행인지
            elif self._is_likely_symbolic_subheader_row(sheet, next_row, max_col):
                should_extend = True
            
            if should_extend:
                header_end = next_row
                extension_count += 1
            else:
                break

        logger.info(f"header range: {header_start}~{header_end}")
        # 2-1) 헤더 후보 행 집합 구성
        header_rows = list(range(header_start, header_end + 1))
        # 2-2) '목차로 되돌아가기' 규칙 적용: 해당 문구가 있는 행과 그 직전 행은 헤더에서 제외
        def row_contains_markers(r: int) -> bool:
            markers = ["목차로 되돌아가기"]
            for c in range(1, max_col + 1):
                val = self._get_merged_top_left_value(sheet, r, c)
                norm = self._normalize_text(val)
                if not norm:
                    continue
                for m in markers:
                    if m in norm:
                        return True
            return False

        marker_rows = {r for r in header_rows if row_contains_markers(r)}
        exclude_rows = set(marker_rows)
        for r in marker_rows:
            if r - 1 >= 1:
                exclude_rows.add(r - 1)
        header_rows = [r for r in header_rows if r not in exclude_rows]
        # 2-3) 모두 제외되었다면, 기준 행이 제외된 경우를 피해서 대체 행 선택
        if not header_rows:
            # 기준 행이 제외되었다면 그 다음 행을 시도, 그렇지 않으면 기준 행 사용
            candidate = base_header_row + 1 if base_header_row in exclude_rows else base_header_row
            # 시트 범위 내 보정
            if candidate > (sheet.max_row or candidate):
                candidate = base_header_row
            header_rows = [candidate]
        # 2-4) 하드코딩 없이 '상징 토큰'으로 구성된 부-헤더 행을 추가로 포함 (최대 2행)
        tail = max(header_rows)
        for r in range(tail + 1, min(tail + 3, (sheet.max_row or tail) + 1)):
            if self._is_likely_symbolic_subheader_row(sheet, r, max_col):
                header_rows.append(r)
            else:
                break
        headers: List[str] = []

        for col in range(1, max_col + 1):
            if self.is_col_hidden(sheet, col):
                continue
            parts: List[str] = []
            for row in header_rows:
                val = self._get_merged_top_left_value(sheet, row, col)
                norm = self._normalize_text(val)
                if norm:
                    if not parts or parts[-1] != norm:
                        parts.append(norm)
            header_name = ' - '.join(parts) if parts else f"Column_{col}"
            headers.append(header_name)

        headers = self._make_unique_headers(headers)
        data_start_row = max(header_rows) + 1
        # 2-5) 헤더 직후 비어있는(또는 숨김) 행들을 건너뛰고 첫 유효 데이터 행으로 보정
        row_probe = data_start_row
        while row_probe <= (sheet.max_row or row_probe):
            # 숨김 행이면 계속 진행
            if self.is_row_hidden(sheet, row_probe):
                row_probe += 1
                continue
            # 내용이 하나도 없으면 계속 진행
            if self._row_non_empty_count(sheet, row_probe, max_col) == 0:
                row_probe += 1
                continue
            break
        data_start_row = row_probe
        logger.info(f"헤더 범위: {min(header_rows)}~{max(header_rows)}행 (기준: {base_header_row}행)")
        logger.info(f"헤더: {headers}")

        return headers, data_start_row, (min(header_rows), max(header_rows))

    def _prepare_col_hidden_cache(self, sheet: Worksheet):
        """
        openpyxl의 column_dimensions를 한 번 훑어 범위(min..max) 단위의 숨김 설정을 캐시한다.
        XML 파싱 없이 dim의 범위 속성(min/max)과 단일 컬럼 속성을 모두 반영한다.
        """
        sheet_name = sheet.title
        if sheet_name in self._sheet_col_hidden_map:
            return
        hidden_map: Dict[int, bool] = {}
        max_col = sheet.max_column or 1
        # 기본값: 모두 False
        for i in range(1, max_col + 1):
            hidden_map[i] = False
        try:
            for key, dim in sheet.column_dimensions.items():
                if dim is None:
                    continue
                # 범위 우선(min..max). openpyxl이 보존한 경우에 한함
                min_idx = getattr(dim, 'min', None)
                max_idx = getattr(dim, 'max', None)
                if isinstance(min_idx, int) and isinstance(max_idx, int) and min_idx >= 1 and max_idx >= min_idx:
                    is_hidden = bool(getattr(dim, 'hidden', False))
                    width_val = getattr(dim, 'width', None)
                    if not is_hidden and (width_val is not None):
                        try:
                            if float(width_val) == 0 or float(width_val) <= 0.2:
                                is_hidden = True
                        except Exception:
                            pass
                    if is_hidden:
                        for ci in range(min_idx, min(max_idx, max_col) + 1):
                            hidden_map[ci] = True
                        logger.debug(f"[col_cache] 범위 {min_idx}-{max_idx} hidden=True 적용")
                    continue
                # 단일 컬럼(키가 문자 인덱스) 처리
                try:
                    col_idx = column_index_from_string(key)
                except Exception:
                    # 키가 문자형이 아닐 수 있음. 이 경우는 스킵
                    continue
                is_hidden = bool(getattr(dim, 'hidden', False))
                width_val = getattr(dim, 'width', None)
                if not is_hidden and (width_val is not None):
                    try:
                        if float(width_val) == 0 or float(width_val) <= 0.2:
                            is_hidden = True
                    except Exception:
                        pass
                # 윤곽 접힘도 보수적으로 반영
                if not is_hidden:
                    outline_level = getattr(dim, 'outlineLevel', 0) or 0
                    collapsed = getattr(dim, 'collapsed', False)
                    if outline_level and collapsed:
                        is_hidden = True
                if is_hidden:
                    hidden_map[col_idx] = True
        except Exception as e:
            logger.debug(f"_prepare_col_hidden_cache 실패(sheet='{sheet_name}'): {e}")
        self._sheet_col_hidden_map[sheet_name] = hidden_map

    def is_col_hidden(self, sheet: Worksheet, col_idx: int) -> bool:
        try:
            # 캐시 우선
            self._prepare_col_hidden_cache(sheet)
            cached = self._sheet_col_hidden_map.get(sheet.title, {}).get(col_idx, False)
            if cached:
                return True
            letter = get_column_letter(col_idx)
            dim = sheet.column_dimensions.get(letter)
            if dim is None:
                # 인접 컬럼의 그룹 접힘(collapse) 영향으로 사실상 비가시일 수 있음 (보수적 처리)
                try:
                    for offset in (-1, 1):
                        adj_idx = col_idx + offset
                        if adj_idx < 1:
                            continue
                        adj_letter = get_column_letter(adj_idx)
                        adj_dim = sheet.column_dimensions.get(adj_letter)
                        if adj_dim and getattr(adj_dim, 'collapsed', False) and getattr(adj_dim, 'outlineLevel', 0):
                            logger.debug(f"[is_col_hidden] 컬럼 {letter}: 인접 {adj_letter} collapsed=True (outlineLevel={getattr(adj_dim,'outlineLevel', None)}) → 숨김 취급")
                            return True
                except Exception as e:
                    logger.debug(f"[is_col_hidden] 컬럼 {letter}: 인접 컬럼 검사 중 오류: {e}")
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: dimension 항목 없음 (sheet='{sheet.title}')")
                return False
            # 명시적 숨김
            if getattr(dim, 'hidden', False):
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: dim.hidden=True 로 숨김 판정")
                return True
            # 너비 0
            if getattr(dim, 'width', None) == 0:
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: dim.width=0 로 숨김 판정")
                return True
            # 보수적 확장: 극소 너비도 숨김으로 간주
            width = getattr(dim, 'width', None)
            if width is not None and isinstance(width, (int, float)) and width <= 0.2:
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: dim.width={width} <= 0.2 (사실상 비가시) → 숨김 판정")
                return True
            # 보수적 확장: 윤곽/그룹 접힘 상태 고려
            outline_level = getattr(dim, 'outlineLevel', 0) or 0
            collapsed = getattr(dim, 'collapsed', False)
            if outline_level and collapsed:
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: outlineLevel={outline_level}, collapsed=True → 숨김 판정")
                return True
            if outline_level:
                # 인접 컬럼이 동일/상위 레벨에서 collapsed인 경우, 현재 컬럼도 접힘에 포함된 것으로 간주
                try:
                    for offset in (-1, 1):
                        adj_idx = col_idx + offset
                        if adj_idx < 1:
                            continue
                        adj_letter = get_column_letter(adj_idx)
                        adj_dim = sheet.column_dimensions.get(adj_letter)
                        if not adj_dim:
                            continue
                        adj_collapsed = getattr(adj_dim, 'collapsed', False)
                        adj_level = (getattr(adj_dim, 'outlineLevel', 0) or 0)
                        if adj_collapsed and adj_level >= outline_level:
                            logger.debug(f"[is_col_hidden] 컬럼 {letter}: 인접 {adj_letter} collapsed=True (adj_level={adj_level} >= {outline_level}) → 숨김 판정")
                            return True
                except Exception as e:
                    logger.debug(f"[is_col_hidden] 컬럼 {letter}: 인접 컬럼 접힘 판정 중 오류: {e}")
            # 숨김 아님: dim 속성 스냅샷 남김
            try:
                attr_names = [a for a in dir(dim) if not a.startswith('_')]
                snapshot = {
                    'hidden': getattr(dim, 'hidden', None),
                    'width': getattr(dim, 'width', None),
                    'bestFit': getattr(dim, 'bestFit', None),
                    'outlineLevel': getattr(dim, 'outlineLevel', None),
                    'collapsed': getattr(dim, 'collapsed', None),
                }
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: 숨김 아님. dim 스냅샷={snapshot}, attrs={attr_names}")
            except Exception as e:
                logger.debug(f"[is_col_hidden] 컬럼 {letter}: dim 속성 로깅 중 오류: {e}")
            return False
        except Exception:
            return False

    def get_headers(self, sheet: Worksheet, header_row: int) -> List[str]:
        """단일 행 기준 폴백 헤더 추출 (호환용)"""
        headers = []
        for cell in sheet[header_row]:
            value = cell.value
            if value is not None:
                headers.append(str(value).strip())
            else:
                headers.append(f"Column_{cell.column}")
        return headers
    
    def _find_column_by_keywords(self, headers: List[str], keywords: List[str]) -> Optional[int]:
        """
        키워드 리스트로 헤더에서 컬럼 인덱스 찾기 (첫 번째만)
        
        Args:
            headers: 헤더 리스트
            keywords: 찾을 키워드 리스트
        
        Returns:
            컬럼 인덱스 (0-based) 또는 None
        """
        for idx, header in enumerate(headers):
            header_normalized = header.strip().replace(' ', '').lower()
            for keyword in keywords:
                keyword_normalized = keyword.strip().replace(' ', '').lower()
                if keyword_normalized in header_normalized:
                    return idx
        return None
    
    def _find_all_columns_by_keywords(self, headers: List[str], keywords: List[str]) -> List[int]:
        """
        키워드 리스트로 헤더에서 모든 매칭 컬럼 인덱스 찾기
        
        Args:
            headers: 헤더 리스트
            keywords: 찾을 키워드 리스트
        
        Returns:
            컬럼 인덱스 리스트 (0-based)
        """
        matching_indices = []
        for idx, header in enumerate(headers):
            header_normalized = header.strip().replace(' ', '').lower()
            for keyword in keywords:
                keyword_normalized = keyword.strip().replace(' ', '').lower()
                if keyword_normalized in header_normalized:
                    matching_indices.append(idx)
                    break  # 하나라도 매칭되면 추가하고 다음 헤더로
        return matching_indices
    
    def detect_sheet_type(self, sheet: Worksheet, sheet_name: str, headers: List[str]) -> SheetType:
        """
        시트 타입 자동 감지
        
        판별 기준:
        1. 목차: 시트명 + 헤더에 목차 키워드
        2. 소프트웨어: 시트명에 "소프트웨어" 포함
        3. 이력관리: 시트명에 "이력" 관련 키워드
        4. REV 관리: 헤더에 "REV" + "WBS" 컬럼 존재
        5. 작성버전 관리: 헤더에 "작성버전" + "관리번호" 컬럼 존재
        6. 첨부파일: 하이퍼링크가 존재하고 REV/작성버전 없음
        
        Args:
            sheet: 워크시트 객체
            sheet_name: 시트 이름
            headers: 헤더 리스트
        
        Returns:
            SheetType enum
        """
        sheet_name_lower = sheet_name.lower()
        
        # 1. 목차 시트 (시트명 + 헤더 키워드)
        for keyword in SHEET_TYPE_KEYWORDS['toc']:
            if keyword.lower() in sheet_name_lower:
                # 헤더에도 목차 관련 키워드가 있는지 확인
                header_text = ' '.join(headers).lower()
                if any(kw.lower() in header_text for kw in SHEET_TYPE_KEYWORDS['toc']):
                    logger.info(f"시트 타입 감지: {sheet_name} → 목차 (시트명+헤더 키워드)")
                    return SheetType.TOC
        
        # 2. 소프트웨어 형상기록 시트 (시트명 우선)
        for keyword in SHEET_TYPE_KEYWORDS['software']:
            if keyword.lower() in sheet_name_lower:
                logger.info(f"시트 타입 감지: {sheet_name} → 소프트웨어 형상기록 (시트명)")
                return SheetType.SOFTWARE
        
        # 3. 이력관리 시트 (시트명)
        for keyword in SHEET_TYPE_KEYWORDS['history']:
            if keyword.lower() in sheet_name_lower:
                logger.info(f"시트 타입 감지: {sheet_name} → 이력관리 (시트명)")
                return SheetType.HISTORY
        
        # 4. REV 관리 문서 (헤더에 REV + WBS 컬럼)
        rev_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['rev'])
        wbs_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['wbs'])
        
        if rev_col_idx is not None and wbs_col_idx is not None:
            logger.info(f"시트 타입 감지: {sheet_name} → REV 관리 (REV+WBS 컬럼)")
            return SheetType.REV_MANAGED
        
        # 5. 작성버전 관리 문서 (헤더에 작성버전 + 관리번호 컬럼)
        version_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['version'])
        manage_no_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['manage_no'])
        
        if version_col_idx is not None and manage_no_col_idx is not None:
            logger.info(f"시트 타입 감지: {sheet_name} → 작성버전 관리 (작성버전+관리번호 컬럼)")
            return SheetType.VERSION_MANAGED
        
        # 6. 첨부파일 시트 (하이퍼링크 존재 여부 확인)
        # 최대 10개 행만 확인 (빠른 판별)
        has_hyperlink = False
        for row_idx in range(1, min(sheet.max_row + 1, 20)):
            for cell in sheet[row_idx]:
                if self.extract_hyperlink(cell):
                    has_hyperlink = True
                    break
            if has_hyperlink:
                break
        
        if has_hyperlink:
            logger.info(f"시트 타입 감지: {sheet_name} → 첨부파일 (하이퍼링크 존재)")
            return SheetType.ATTACHMENT
        
        # 기본값: 미분류
        logger.warning(f"시트 타입 감지: {sheet_name} → 미분류 (알 수 없는 타입)")
        return SheetType.UNKNOWN
    
    def is_row_hidden(self, sheet: Worksheet, row_idx: int) -> bool:
        """행이 숨겨져 있거나 높이가 0인지 확인"""
        try:
            row_dimension = sheet.row_dimensions[row_idx]
            # 숨김 처리되었거나 높이가 0인 경우
            if row_dimension.hidden:
                return True
            if row_dimension.height is not None and row_dimension.height == 0:
                return True
            return False
        except:
            return False
    
    def extract_hyperlink(self, cell: Cell) -> Optional[str]:
        """셀에서 하이퍼링크 추출"""
        if cell.hyperlink:
            # 하이퍼링크 객체가 있는 경우
            target = cell.hyperlink.target
            if target:
                return target
        
        # 수식에서 하이퍼링크 추출 시도
        if cell.value and isinstance(cell.value, str):
            if cell.value.startswith('=HYPERLINK'):
                # =HYPERLINK("url", "display") 형식 파싱
                try:
                    start = cell.value.index('"') + 1
                    end = cell.value.index('"', start)
                    return cell.value[start:end]
                except:
                    pass
        
        return None
    
    def generate_document_key(
        self, 
        sheet_type: SheetType, 
        sheet_name: str, 
        row_data: Dict[str, str],
        headers: List[str]
    ) -> Optional[str]:
        """
        문서 고유 키 생성 (Revision 관리용)
        
        Args:
            sheet_type: 시트 타입
            sheet_name: 시트 이름
            row_data: 행 데이터 (헤더: 값)
            headers: 헤더 리스트
        
        Returns:
            문서 키 또는 None
        """
        if sheet_type == SheetType.REV_MANAGED:
            # REV 관리: 모든 WBS 컬럼 값 합치기 + 시트명
            wbs_col_indices = self._find_all_columns_by_keywords(headers, COLUMN_NAME_MAPPINGS['wbs'])
            
            if wbs_col_indices:
                wbs_values = []
                for idx in wbs_col_indices:
                    if idx < len(headers):
                        wbs_header = headers[idx]
                        wbs_value = row_data.get(wbs_header, '').strip()
                        if wbs_value:
                            wbs_values.append(wbs_value)
                
                if wbs_values:
                    # 모든 WBS 값을 '-'로 연결
                    combined_wbs = '-'.join(wbs_values)
                    # 공백, 특수문자 정규화
                    key = f"{combined_wbs}_{sheet_name}"
                    key = key.replace(' ', '_').replace('/', '_').replace('\\', '_')
                    logger.debug(f"생성된 문서 키 (WBS 컬럼 {len(wbs_values)}개): {key}")
                    return key
        
        elif sheet_type == SheetType.VERSION_MANAGED:
            # 작성버전 관리: 관리번호 + 시트명
            manage_no_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['manage_no'])
            if manage_no_col_idx is not None and manage_no_col_idx < len(headers):
                manage_no_header = headers[manage_no_col_idx]
                manage_no_value = row_data.get(manage_no_header, '').strip()
                if manage_no_value:
                    # 공백, 특수문자 정규화
                    key = f"{manage_no_value}_{sheet_name}"
                    key = key.replace(' ', '_').replace('/', '_').replace('\\', '_')
                    return key
        
        return None
    
    def get_revision_value(
        self, 
        sheet_type: SheetType, 
        row_data: Dict[str, str],
        headers: List[str],
        row_cells: List[Cell] = None
    ) -> Optional[str]:
        """
        행에서 revision 값 추출
        
        Args:
            sheet_type: 시트 타입
            row_data: 행 데이터 (헤더: 값)
            headers: 헤더 리스트
            row_cells: 행의 셀 리스트 (하이퍼링크 텍스트 추출용, 선택)
        
        Returns:
            revision 값 또는 None
            
        Note:
            - REV 관리: rev 컬럼의 하이퍼링크 텍스트가 revision 명
            - 작성버전 관리: 관리번호 컬럼의 하이퍼링크 텍스트가 revision 명
        """
        if sheet_type == SheetType.REV_MANAGED:
            # REV 컬럼에서 값 추출 (하이퍼링크 텍스트 우선)
            rev_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['rev'])
            if rev_col_idx is not None and rev_col_idx < len(headers):
                # 1. 셀이 제공된 경우: 하이퍼링크 텍스트(셀 표시 텍스트) 추출
                if row_cells and rev_col_idx < len(row_cells):
                    cell = row_cells[rev_col_idx]
                    if cell.value is not None:
                        rev_value = str(cell.value).strip()
                        if rev_value:
                            logger.debug(f"REV 컬럼에서 하이퍼링크 텍스트 추출: {rev_value}")
                            return rev_value
                
                # 2. 폴백: row_data에서 추출
                rev_header = headers[rev_col_idx]
                rev_value = row_data.get(rev_header, '').strip()
                return rev_value if rev_value else None
        
        elif sheet_type == SheetType.VERSION_MANAGED:
            # 작성버전 컬럼에서 값 추출
            version_col_idx = self._find_column_by_keywords(headers, COLUMN_NAME_MAPPINGS['version'])
            if version_col_idx is not None and version_col_idx < len(headers):
                # 1. 셀이 제공된 경우: 하이퍼링크 텍스트(셀 표시 텍스트) 추출
                if row_cells and version_col_idx < len(row_cells):
                    cell = row_cells[version_col_idx]
                    if cell.value is not None:
                        version_value = str(cell.value).strip()
                        if version_value:
                            logger.debug(f"작성버전 컬럼에서 하이퍼링크 텍스트 추출: {version_value}")
                            return version_value
                
                # 2. 폴백: row_data에서 추출
                version_header = headers[version_col_idx]
                version_value = row_data.get(version_header, '').strip()
                return version_value if version_value else None
        
        return None
    
    def extract_sheet_as_excel(self, sheet_name: str, output_dir: Path) -> Optional[Path]:
        """
        특정 시트를 독립된 Excel 파일로 추출
        
        Args:
            sheet_name: 추출할 시트 이름
            output_dir: 출력 디렉토리
        
        Returns:
            추출된 Excel 파일 경로 또는 None
        """
        if not self.workbook:
            logger.error("워크북이 로드되지 않았습니다.")
            return None
        
        try:
            # 새 워크북 생성
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # 기본 시트 제거
            
            # 원본 시트 복사
            source_sheet = self.workbook[sheet_name]
            target_sheet = new_wb.create_sheet(sheet_name)
            
            # 모든 셀 복사 (값, 스타일 포함)
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value
                    
                    # 스타일 복사 (간단 버전)
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()
            
            # 컬럼 너비 복사
            for col_letter in source_sheet.column_dimensions:
                if col_letter in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col_letter].width = \
                        source_sheet.column_dimensions[col_letter].width
            
            # 행 높이 복사
            for row_num in source_sheet.row_dimensions:
                if row_num in source_sheet.row_dimensions:
                    target_sheet.row_dimensions[row_num].height = \
                        source_sheet.row_dimensions[row_num].height
            
            # 파일 저장
            output_path = output_dir / f"{sheet_name}.xlsx"
            new_wb.save(output_path)
            new_wb.close()
            
            logger.info(f"시트 '{sheet_name}'를 Excel 파일로 추출: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"시트 '{sheet_name}' Excel 추출 실패: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def convert_sheet_to_text_chunks(
        self, 
        sheet_name: str, 
        max_length: int = MAX_TEXT_LENGTH,
        return_rows_as_list: bool = False
    ) -> List[Any]:
        """
        시트 전체를 텍스트로 변환 (이력관리/소프트웨어 형상기록용)
        Row 단위로 체크하여 max_length를 초과하면 여러 청크로 분할
        
        형식:
        헤더1: 값1
        헤더2: 값2
        ---
        헤더1: 값3
        헤더2: 값4
        ...
        
        Args:
            sheet_name: 시트 이름
            max_length: 청크당 최대 문자 수 (토큰 제한)
            return_rows_as_list: True면 청크를 문자열이 아닌 Row 리스트로 반환
        
        Returns:
            변환된 텍스트 청크 리스트 (str 또는 List[str])
        """
        if not self.workbook:
            logger.error("워크북이 로드되지 않았습니다.")
            return []
        
        sheet = self.workbook[sheet_name]
        
        # 헤더 생성 (병합/다단 대응)
        headers, data_start_row, header_span = self.build_headers_and_data_start(sheet)
        
        # === 첫 컬럼 기준 그룹핑(+최대 5행 버퍼) 로직을 적용하여 텍스트 생성 ===
        def first_col_has_value(cells: List[Cell]) -> bool:
            if not cells:
                return False
            v = cells[0].value
            return v is not None and str(v).strip() != ''
        
        def merge_metadata(dst: Dict[str, str], src: Dict[str, str]):
            for k, v in src.items():
                if not v:
                    continue
                if k not in dst or not dst[k]:
                    dst[k] = v
                else:
                    if v not in dst[k]:
                        dst[k] = f"{dst[k]} / {v}"
        
        def make_text_from_metadata(metadata: Dict[str, str]) -> Optional[str]:
            if not metadata:
                return None
            parts: List[str] = []
            
            # 시트명 정보 추가
            parts.append(f"시트명: {sheet_name}")
            
            # 헤더 순서대로 출력 (가시 컬럼만)
            for header in headers:
                val = metadata.get(header, '')
                if val:
                    parts.append(f"{header}: {val}")
            return '\n'.join(parts) if parts else None
        
        chunks: List[Any] = []
        current_chunk_rows: List[str] = []
        current_length: int = 0
        
        current_metadata: Optional[Dict[str, str]] = None
        pending_rows: List[Dict[str, str]] = []  # 병합 전 대기 메타데이터 (최대 5행)
        
        def flush_current_to_chunks():
            nonlocal current_metadata, current_chunk_rows, current_length
            if not current_metadata:
                return
            row_text = make_text_from_metadata(current_metadata)
            current_metadata = None
            if not row_text:
                return
            row_length = len(row_text) + len(ROW_SEPARATOR)
            if current_length + row_length > max_length and current_chunk_rows:
                # 청크 마감
                if return_rows_as_list:
                    chunks.append(list(current_chunk_rows))
                else:
                    chunk_text = ROW_SEPARATOR.join(current_chunk_rows)
                    chunks.append(chunk_text)
                
                logger.debug(f"청크 {len(chunks)} 완료: {len(current_chunk_rows)}개 행")
                # 새 청크 시작
                current_chunk_rows = [row_text]
                current_length = row_length
            else:
                current_chunk_rows.append(row_text)
                current_length += row_length
        
        for row_idx in range(data_start_row, sheet.max_row + 1):
            # 숨김 행 제외
            if self.is_row_hidden(sheet, row_idx):
                continue
            
            row_cells = list(sheet[row_idx])
            # 빈 행 건너뛰기
            if all(cell.value is None for cell in row_cells):
                continue
            
            # 현재 행의 메타데이터 구성 (병합영역 좌상단 값 사용, 숨김 컬럼 제외)
            row_metadata: Dict[str, str] = {}
            header_idx = 0
            for col_idx, _cell in enumerate(row_cells):
                col_number = col_idx + 1
                if self.is_col_hidden(sheet, col_number):
                    continue
                if header_idx < len(headers):
                    header = headers[header_idx]
                    header_idx += 1
                    # 수식 셀은 계산된 값(data_only)을 우선 사용
                    merged_val = self._get_merged_top_left_value_evaluated(sheet, row_idx, col_number)
                    text = self._normalize_text(merged_val)
                    if text:
                        row_metadata[header] = text
            
            if first_col_has_value(row_cells):
                # 기존 레코드가 있으면 플러시
                flush_current_to_chunks()
                # 새 레코드 시작
                current_metadata = dict(row_metadata)
                # 대기(pending)된 선행 행들 병합 (최대 5행 누적)
                if pending_rows:
                    for pm in pending_rows:
                        merge_metadata(current_metadata, pm)
                    pending_rows.clear()
                continue
            
            # 첫 컬럼이 비었을 때: 현재 레코드에 병합, 없으면 대기열에 추가
            if current_metadata is not None:
                merge_metadata(current_metadata, row_metadata)
            else:
                if len(pending_rows) < 7:
                    pending_rows.append(dict(row_metadata))
                else:
                    pending_rows.pop(0)
                    pending_rows.append(dict(row_metadata))
        
        # 마지막 레코드 플러시
        flush_current_to_chunks()
        
        # 마지막 청크 추가
        if current_chunk_rows:
            if return_rows_as_list:
                chunks.append(list(current_chunk_rows))
            else:
                chunk_text = ROW_SEPARATOR.join(current_chunk_rows)
                chunks.append(chunk_text)
            logger.debug(f"청크 {len(chunks)} 완료: {len(current_chunk_rows)}개 행")
        
        logger.info(f"시트 '{sheet_name}' 텍스트 변환 완료: 총 {len(chunks)}개 청크")
        return chunks
    
    def process_sheet(self, sheet_name: str, early_stop_no_value: Optional[int] = None) -> Tuple[SheetType, List[Dict], List[str]]:
        """
        시트 처리 - 시트 타입 감지, 하이퍼링크와 메타데이터 추출
        
        Returns:
            Tuple[SheetType, List[Dict], List[str]]: (
                시트 타입,
                [
                    {
                        'hyperlink': 'file_url',
                        'metadata': {'col1': 'val1', 'col2': 'val2', ...},
                        'row_number': 5,
                        'document_key': 'WBS-1.2.3_시트명' (revision 관리 시트만),
                        'revision': '1.0' (revision 관리 시트만)
                    },
                    ...
                ],
                헤더 리스트
            )
        """
        if not self.workbook:
            logger.error("워크북이 로드되지 않았습니다.")
            return SheetType.UNKNOWN, [], []
        
        sheet = self.workbook[sheet_name]
        logger.log_sheet_start(sheet_name)
        
        # 헤더 생성 (병합/다단 대응)
        headers, data_start_row, header_span = self.build_headers_and_data_start(sheet)
        
        # 시트 타입 감지
        sheet_type = self.detect_sheet_type(sheet, sheet_name, headers)
        
        results: List[Dict] = []
        
        # 연속 행 병합 로직 (첫 컬럼 우선 + 5행 버퍼)
        def count_non_empty(cells: List[Cell]) -> int:
            return sum(1 for c in cells if c.value is not None and str(c.value).strip())

        def first_col_has_value(cells: List[Cell]) -> bool:
            if not cells:
                return False
            v = cells[0].value
            return v is not None and str(v).strip() != ''
 
        def merge_metadata(dst: Dict[str, str], src: Dict[str, str]):
            for k, v in src.items():
                if not v:
                    continue
                if k not in dst or not dst[k]:
                    dst[k] = v
                else:
                    if v not in dst[k]:
                        dst[k] = f"{dst[k]} / {v}"
 
        current: Optional[Dict] = None
        pending_rows: List[Tuple[Dict[str, str], List[str], int]] = []  # (metadata, hyperlinks, row_idx)

        def finalize_current():
            if not current:
                return
            # 하위 호환: hyperlinks가 있고 hyperlink가 비어있으면 첫 항목을 대표 링크로 설정
            if current.get('hyperlinks') and not current.get('hyperlink'):
                if isinstance(current.get('hyperlinks'), list) and current['hyperlinks']:
                    current['hyperlink'] = current['hyperlinks'][0]
            if sheet_type in [SheetType.ATTACHMENT, SheetType.REV_MANAGED, SheetType.VERSION_MANAGED] and not (current.get('hyperlink') or (isinstance(current.get('hyperlinks'), list) and current.get('hyperlinks'))):
                return
            if sheet_type in [SheetType.REV_MANAGED, SheetType.VERSION_MANAGED]:
                metadata = current['metadata']
                document_key = self.generate_document_key(sheet_type, sheet_name, metadata, headers)
                revision = self.get_revision_value(sheet_type, metadata, headers, None)
                if document_key:
                    current['document_key'] = document_key
                if revision:
                    current['revision'] = revision
            results.append(current.copy())

        # 데이터 행 처리 (그룹핑 적용)
        no_value_streak = 0
        for row_idx in range(data_start_row, sheet.max_row + 1):
            # 테스트 모드: 행 수 제한 확인
            if TEST_MODE and TEST_MAX_ROWS > 0 and len(results) >= TEST_MAX_ROWS:
                logger.warning(f"[테스트 모드] 시트 '{sheet_name}': {TEST_MAX_ROWS}개 행 제한 도달, 나머지 행 건너뜁니다.")
                break
            
            # 숨겨진 행 또는 높이가 0인 행 제외
            if self.is_row_hidden(sheet, row_idx):
                logger.debug(f"{row_idx}행은 숨김 처리되었거나 높이가 0이어서 건너뜁니다.")
                continue
            
            row_cells = list(sheet[row_idx])
            
            # 빈 행 건너뛰기
            if all(cell.value is None for cell in row_cells):
                continue
            
            # 하이퍼링크 찾기 (병합영역 좌상단 포함) - 행 단위로 모두 수집
            hyperlinks_in_row: List[str] = []
            for col_i, cell in enumerate(row_cells, start=1):
                link = self.extract_hyperlink(cell)
                if not link:
                    # 병합영역의 좌상단에서 재시도
                    coord = cell.coordinate
                    for mrange in sheet.merged_cells.ranges:
                        if coord in mrange:
                            top_left_cell = sheet.cell(row=mrange.min_row, column=mrange.min_col)
                            link = self.extract_hyperlink(top_left_cell)
                            break
                if link:
                    if link not in hyperlinks_in_row:
                        hyperlinks_in_row.append(link)
            
            # 메타데이터 구성 (병합영역 좌상단 값 사용)
            row_metadata: Dict[str, str] = {}
            header_idx = 0
            for col_idx, cell in enumerate(row_cells):
                col_number = col_idx + 1
                if self.is_col_hidden(sheet, col_number):
                    continue
                if header_idx < len(headers):
                    header = headers[header_idx]
                    header_idx += 1
                    merged_val = self._get_merged_top_left_value_evaluated(sheet, row_idx, col_number)
                    text = self._normalize_text(merged_val)
                    if text:
                        row_metadata[header] = text

            # 첫 컬럼 기준 그룹핑
            if first_col_has_value(row_cells):
                # 조기 종료 카운터 리셋
                no_value_streak = 0
                # 기존 레코드 마감
                finalize_current()

                # 새 레코드 시작
                current = {
                    'hyperlink': hyperlinks_in_row[0] if hyperlinks_in_row else None,
                    'hyperlinks': list(hyperlinks_in_row) if hyperlinks_in_row else [],
                    'metadata': row_metadata,
                    'row_number': row_idx,
                    'sheet_name': sheet_name
                }

                # 대기(pending)된 선행 행들 병합 (최대 5행 누적)
                if pending_rows:
                    for pm, phs, pi in pending_rows:
                        merge_metadata(current['metadata'], pm)
                        # 링크 병합 (대표 링크 부재 시 대표 링크 설정)
                        if phs:
                            if not current.get('hyperlinks'):
                                current['hyperlinks'] = []
                            for l in phs:
                                if l not in current['hyperlinks']:
                                    current['hyperlinks'].append(l)
                            if not current.get('hyperlink') and current['hyperlinks']:
                                current['hyperlink'] = current['hyperlinks'][0]
                        current['row_number'] = pi
                    pending_rows.clear()
                continue

            # 첫 컬럼 비어있음: 현재 레코드에 병합하거나, 없으면 대기열에 추가(최대 5행)
            if current is not None:
                merge_metadata(current['metadata'], row_metadata)
                # 링크 병합
                if hyperlinks_in_row:
                    if not current.get('hyperlinks'):
                        current['hyperlinks'] = []
                    for l in hyperlinks_in_row:
                        if l not in current['hyperlinks']:
                            current['hyperlinks'].append(l)
                    if not current.get('hyperlink'):
                        current['hyperlink'] = current['hyperlinks'][0]
                current['row_number'] = row_idx
            else:
                if len(pending_rows) < 7:
                    pending_rows.append((row_metadata, list(hyperlinks_in_row), row_idx))
                else:
                    # 버퍼 초과 시 가장 오래된 항목은 폐기하여 메모리 제한 유지
                    pending_rows.pop(0)
                    pending_rows.append((row_metadata, list(hyperlinks_in_row), row_idx))

            # 조기 종료 카운팅 (유효 값이 전혀 없을 때만 증가)
            if row_metadata:
                no_value_streak = 0
            else:
                no_value_streak += 1
                if early_stop_no_value is not None and no_value_streak >= early_stop_no_value:
                    logger.info(f"연속 {early_stop_no_value}개 무값 행 감지 → 스캔 조기 종료 (row={row_idx})")
                    break

        # 마지막 레코드 마감
        finalize_current()
        
        logger.log_sheet_end(sheet_name, len(results))
        return sheet_type, results, headers
    
    def process_all_sheets(self) -> Dict[str, Tuple[SheetType, List[Dict], List[str]]]:
        """
        모든 시트 처리 (숨겨진 시트 제외)
        
        Returns:
            Dict[str, Tuple[SheetType, List[Dict], List[str]]]: {
                'sheet1': (시트타입, [항목들...], [헤더들...]),
                'sheet2': (시트타입, [항목들...], [헤더들...]),
                ...
            }
        """
        if not self.load_workbook():
            return {}
        
        all_results = {}
        sheet_names = self.get_sheet_names()
        
        logger.info(f"총 {len(sheet_names)}개 시트 발견: {sheet_names}")
        
        # 테스트 모드 알림
        if TEST_MODE:
            logger.warning(f"[테스트 모드] 활성화됨 - 최대 {TEST_MAX_SHEETS}개 시트, 시트당 {TEST_MAX_ROWS}개 행만 처리")
        
        processed_sheet_count = 0
        for sheet_name in sheet_names:
            # 테스트 모드: 시트 수 제한 확인
            if TEST_MODE and TEST_MAX_SHEETS > 0 and processed_sheet_count >= TEST_MAX_SHEETS:
                logger.warning(f"[테스트 모드] {TEST_MAX_SHEETS}개 시트 제한 도달, 나머지 시트 건너뜀")
                break
            
            try:
                sheet = self.workbook[sheet_name]
                
                # 시트가 숨겨져 있는지 확인
                if sheet.sheet_state == 'hidden' or sheet.sheet_state == 'veryHidden':
                    logger.info(f"시트 '{sheet_name}'는 숨김 처리되어 건너뜁니다.")
                    continue
                
                sheet_type, results, headers = self.process_sheet(sheet_name)
                all_results[sheet_name] = (sheet_type, results, headers)
                processed_sheet_count += 1
            except Exception as e:
                logger.error(f"시트 '{sheet_name}' 처리 중 오류: {e}")
                import traceback
                logger.error(traceback.format_exc())
                continue
        
        return all_results
    
    def close(self):
        """워크북 닫기"""
        if self.workbook:
            self.workbook.close()
            logger.info("워크북 닫기 완료")

