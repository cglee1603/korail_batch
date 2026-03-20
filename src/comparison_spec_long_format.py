"""
고속차량 제원 비교표 형태의 넓은 비교 엑셀을 롱 포맷(표준 테이블)으로 변환한다.

- 파일명이 설정된 접두어로 시작하는 .xlsx/.xlsm 만 대상 (filesystem_batch → FileHandler 경로).
- 기대 레이아웃(가장 흔한 형태): 1행 제목(선택) → 헤더 행에 열1「구분」, 열2 비어 있거나 항목명,
  열3~ 에 차량(열) 헤더. 데이터는 열1=대분류(세로 병합), 열2=세부항목, 열3~=값.
- 출력: 단순 6열 테이블(xlsx 기본, 선택 시 csv 동시 생성) → RAGFlow table 파서에 적합.
"""
from __future__ import annotations

import csv
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from logger import logger

# ----- 설정 (환경변수, 하드코딩 최소화) -----
def _env_str(key: str, default: str) -> str:
    v = os.getenv(key)
    return (v.strip() if v else default) if v is not None else default


def _env_bool(key: str, default: bool) -> bool:
    v = os.getenv(key)
    if v is None:
        return default
    return v.strip().lower() in ("1", "true", "yes", "y", "on")


def filename_matches_comparison_spec(path: Path) -> bool:
    """변환 대상 파일명인지 (stem 기준 접두어)."""
    prefix = _env_str(
        "COMPARISON_SPEC_FILENAME_PREFIX",
        "고속차량 제원 비교표",
    )
    return path.stem.startswith(prefix)


def _norm(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).replace("\r\n", "\n").replace("\r", "\n").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_display(sheet: Worksheet, row: int, col: int) -> str:
    """병합 시 좌상단 값, 단 현재 셀에 값이 있으면 우선."""
    cell = sheet.cell(row=row, column=col)
    if cell.value is not None and str(cell.value).strip():
        return _norm(cell.value)
    coord = cell.coordinate
    for mrange in sheet.merged_cells.ranges:
        if coord in mrange:
            tl = sheet.cell(row=mrange.min_row, column=mrange.min_col)
            return _norm(tl.value)
    return ""


def _is_row_hidden(sheet: Worksheet, row: int) -> bool:
    try:
        dim = sheet.row_dimensions.get(row)
        return bool(dim and dim.hidden)
    except Exception:
        return False


def _is_col_hidden(sheet: Worksheet, col: int) -> bool:
    try:
        for key, dim in sheet.column_dimensions.items():
            if dim is None:
                continue
            mn = getattr(dim, "min", None)
            mx = getattr(dim, "max", None)
            if isinstance(mn, int) and isinstance(mx, int) and mn <= col <= mx:
                if getattr(dim, "hidden", False):
                    return True
            elif key and key.isalpha():
                from openpyxl.utils import column_index_from_string

                try:
                    if column_index_from_string(key) == col and getattr(dim, "hidden", False):
                        return True
                except Exception:
                    pass
    except Exception:
        pass
    return False


def _sheet_document_title(sheet: Worksheet, filename_stem: str) -> str:
    """시트 상단 병합 제목 또는 첫 행 텍스트."""
    max_col = min(sheet.max_column or 1, 30)
    for r in range(1, min(6, (sheet.max_row or 1) + 1)):
        for mrange in sheet.merged_cells.ranges:
            if mrange.min_row == mrange.max_row == r and (mrange.max_col - mrange.min_col + 1) >= max(
                3, max_col // 2
            ):
                tl = sheet.cell(row=mrange.min_row, column=mrange.min_col)
                t = _norm(tl.value)
                if t and len(t) >= 4:
                    return t
        row_text = " ".join(_cell_display(sheet, r, c) for c in range(1, max_col + 1))
        row_text = _norm(row_text)
        if len(row_text) >= 6 and "구분" not in row_text[:20]:
            return row_text[:200]
    return filename_stem


# 열2에 오는 항목명 헤더(차량 열이 아님) — 실무 서식「구분 | 상세내역 | KTX | …」대응
_META_HEADER_TOKENS = frozenset(
    (
        "구분",
        "항목",
        "세부항목",
        "세부 항목",
        "상세내역",
        "상세 내역",
        "분류",
        "종별",
        "내용",
        "비고",
        "세부",
        "",
    )
)


def _detect_header_layout(
    sheet: Worksheet,
) -> Tuple[Optional[int], int, int]:
    """
    Returns:
        (header_row_1based, vehicle_start_col_1based, item_col_1based)
        실패 시 (None, 0, 0)
    """
    max_scan_row = min(35, (sheet.max_row or 0) + 1)
    max_col = min(sheet.max_column or 26, 200)

    for hr in range(1, max_scan_row):
        if _is_row_hidden(sheet, hr):
            continue
        c1 = _cell_display(sheet, hr, 1)
        if "구분" not in c1:
            continue

        b1 = _cell_display(sheet, hr, 2)
        vehicle_cols: List[int] = []
        for c in range(2, max_col + 1):
            if _is_col_hidden(sheet, c):
                continue
            h = _cell_display(sheet, hr, c)
            if not h or h in _META_HEADER_TOKENS:
                continue
            vehicle_cols.append(c)

        if len(vehicle_cols) < 2:
            continue

        first_v = vehicle_cols[0]
        if first_v >= 3:
            # A=구분, B=항목열(헤더 비어 있음), C~=차량
            return hr, first_v, 2
        # B열부터 차량명
        return hr, 2, 1

    return None, 0, 0


def _unique_vehicle_headers(sheet: Worksheet, hr: int, vehicle_start: int, max_col: int) -> Dict[int, str]:
    seen: Dict[str, int] = {}
    out: Dict[int, str] = {}
    for c in range(vehicle_start, max_col + 1):
        if _is_col_hidden(sheet, c):
            continue
        h = _cell_display(sheet, hr, c)
        if not h:
            continue
        base = h
        if base not in seen:
            seen[base] = 1
            out[c] = base
        else:
            seen[base] += 1
            out[c] = f"{base} ({seen[base]})"
    return out


def _row_has_data_values(sheet: Worksheet, row: int, vehicle_start: int, max_col: int) -> bool:
    for c in range(vehicle_start, max_col + 1):
        if _is_col_hidden(sheet, c):
            continue
        if _cell_display(sheet, row, c):
            return True
    return False


def extract_long_rows_for_sheet(
    sheet: Worksheet,
    sheet_name: str,
    doc_title: str,
) -> List[Dict[str, str]]:
    hr, vehicle_start, item_col = _detect_header_layout(sheet)
    if hr is None:
        return []

    max_col = sheet.max_column or vehicle_start
    vmap = _unique_vehicle_headers(sheet, hr, vehicle_start, max_col)
    if len(vmap) < 2:
        return []

    rows_out: List[Dict[str, str]] = []
    current_group = ""

    for r in range(hr + 1, (sheet.max_row or 0) + 1):
        if _is_row_hidden(sheet, r):
            continue
        if not _row_has_data_values(sheet, r, vehicle_start, max_col):
            continue

        g = _cell_display(sheet, r, 1)
        if item_col == 2:
            if g:
                current_group = g
            item = _cell_display(sheet, r, 2)
        else:
            # 구분|KTX|… (열2부터 차량): 열1=세부항목만 사용 (대분류 열 없음)
            item = g
            current_group = ""

        for col_idx, vehicle_name in sorted(vmap.items()):
            val = _cell_display(sheet, r, col_idx)
            if not val:
                continue
            rows_out.append(
                {
                    "표제목": doc_title,
                    "시트명": sheet_name,
                    "대분류": current_group,
                    "세부항목": item,
                    "차량종류": vehicle_name,
                    "값": val,
                }
            )

    return rows_out


def _write_xlsx(path: Path, records: List[Dict[str, str]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "spec_long"
    headers = ["표제목", "시트명", "대분류", "세부항목", "차량종류", "값"]
    ws.append(headers)
    for rec in records:
        ws.append([rec.get(h, "") for h in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_csv(path: Path, records: List[Dict[str, str]]) -> None:
    headers = ["표제목", "시트명", "대분류", "세부항목", "차량종류", "값"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(records)


def transform_comparison_spec_workbook(
    source_path: Path,
    output_dir: Path,
) -> Optional[Path]:
    """
    변환 성공 시 업로드용 주 경로 반환 (기본: xlsx).
    csv 동시 생성 시에도 업로드 파일은 xlsx 유지 (filesystem이 table 파서 연동).
    """
    if source_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return None

    also_csv = _env_bool("COMPARISON_SPEC_ALSO_EXPORT_CSV", False)
    try:
        wb = openpyxl.load_workbook(source_path, data_only=True, read_only=False)
    except Exception as e:
        logger.warning(f"[comparison_spec_long] 워크북 로드 실패: {source_path.name} — {e}")
        return None

    all_recs: List[Dict[str, str]] = []
    stem = source_path.stem

    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            title = _sheet_document_title(ws, stem)
            part = extract_long_rows_for_sheet(ws, sn, title)
            all_recs.extend(part)
    finally:
        wb.close()

    if not all_recs:
        logger.info(
            f"[comparison_spec_long] 롱 포맷 추출 행 없음 (레이아웃 불일치 가능): {source_path.name}"
        )
        return None

    output_dir = Path(output_dir)
    out_xlsx = output_dir / f"{stem}_long_table.xlsx"
    out_csv = output_dir / f"{stem}_long_table.csv"

    _write_xlsx(out_xlsx, all_recs)
    logger.info(
        f"[comparison_spec_long] 롱 포맷 변환: {source_path.name} → {out_xlsx.name} ({len(all_recs)}행)"
    )

    if also_csv:
        _write_csv(out_csv, all_recs)
        logger.info(f"[comparison_spec_long] CSV 부가 저장: {out_csv.name}")

    return out_xlsx


def transform_comparison_spec_if_applicable(
    file_path: Path,
    output_dir: Path,
) -> Optional[Path]:
    """
    파일명 접두어가 맞을 때만 변환 시도. 실패·비적용 시 None → 호출측에서 일반 Excel 단순화.
    """
    if not filename_matches_comparison_spec(file_path):
        return None
    return transform_comparison_spec_workbook(file_path, output_dir)
